import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from src.database import init_db, get_scored_cases, get_latest_decision_per_transaction

st.set_page_config(page_title="Export | Fraud Analyzer", page_icon="📥", layout="wide")
init_db()

TIER_FILL_COLORS = {
    "Critical": "FFEF4444",
    "High": "FFFB923C",
    "Medium": "FFFBBF24",
    "Low": "FF4ADE80"
}

st.title("📥 Export Results")
st.markdown("Download reviewed cases as a color-coded Excel report.")

scored = get_scored_cases()

if scored.empty:
    st.warning("No scored transactions. Go to **2 Score** first.")
    st.stop()

decisions = get_latest_decision_per_transaction()
if not decisions.empty:
    scored = scored.merge(
        decisions[["transaction_id", "decision", "notes", "decided_at"]],
        on="transaction_id",
        how="left"
    )
    scored["decision"] = scored["decision"].fillna("pending")
    scored["notes"] = scored["notes"].fillna("")
    scored["decided_at"] = scored["decided_at"].fillna("")
else:
    scored["decision"] = "pending"
    scored["notes"] = ""
    scored["decided_at"] = ""

st.subheader("Export Filters")

col1, col2 = st.columns(2)
include_tiers = col1.multiselect(
    "Include Risk Tiers",
    ["Critical", "High", "Medium", "Low"],
    default=["Critical", "High", "Medium"]
)
include_statuses = col2.multiselect(
    "Include Decision Status",
    ["pending", "escalated", "reviewed", "cleared"],
    default=["pending", "escalated", "reviewed", "cleared"]
)

export_df = scored[
    scored["risk_tier"].isin(include_tiers) &
    scored["decision"].isin(include_statuses)
].copy()

st.markdown("---")
col1, col2, col3 = st.columns(3)
col1.metric("Rows to Export", len(export_df))
col2.metric("Escalated", len(export_df[export_df["decision"] == "escalated"]))
col3.metric("Pending Review", len(export_df[export_df["decision"] == "pending"]))

if not export_df.empty:
    st.subheader("Export Preview (first 20 rows)")
    st.dataframe(
        export_df[["date", "merchant", "category", "amount", "location",
                   "final_score", "risk_tier", "decision", "notes"]].head(20),
        width="stretch"
    )
else:
    st.info("No rows match the selected filters.")


def build_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()

    export_cols = ["date", "merchant", "category", "amount", "card_last4", "location",
                   "rule_score", "ml_score", "final_score", "risk_tier", "decision", "notes", "decided_at"]
    available_cols = [c for c in export_cols if c in df.columns]
    clean_df = df[available_cols].copy()
    clean_df.columns = [c.replace("_", " ").title() for c in clean_df.columns]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        clean_df.to_excel(writer, index=False, sheet_name="Fraud Cases")
        ws = writer.sheets["Fraud Cases"]

        header_fill = PatternFill(start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        tier_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value and "Risk Tier" in str(cell.value):
                tier_col_idx = idx

        if tier_col_idx:
            for row in ws.iter_rows(min_row=2):
                tier_val = row[tier_col_idx - 1].value
                fill_color = TIER_FILL_COLORS.get(tier_val, "FFFFFFFF")
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                for cell in row:
                    cell.fill = fill

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    return output.getvalue()


st.markdown("---")
if not export_df.empty:
    if st.button("📥 Generate Excel Report", type="primary"):
        with st.spinner("Building Excel report..."):
            st.session_state["excel_bytes"] = build_excel(export_df)
            st.session_state["excel_row_count"] = len(export_df)

    if "excel_bytes" in st.session_state:
        st.download_button(
            label=f"⬇️ Download fraud_analysis_report.xlsx ({st.session_state.get('excel_row_count', 0)} rows)",
            data=st.session_state["excel_bytes"],
            file_name="fraud_analysis_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        st.success("✅ Report ready — click the download button above.")
else:
    st.button("📥 Generate Excel Report", type="primary", disabled=True)
